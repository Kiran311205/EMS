from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.http import HttpResponse
from apps.accounts.permissions import IsHR
from apps.employees.models import Employee
from apps.salary.models import SalaryRecord
from apps.assets.models import Asset
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import datetime


def make_header_style():
    font = Font(bold=True, color="FFFFFF", size=11)
    fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    align = Alignment(horizontal="center", vertical="center")
    return font, fill, align


class EmployeeReportView(APIView):
    permission_classes = [IsAuthenticated, IsHR]

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Employees"

        font, fill, align = make_header_style()
        headers = ['Emp ID', 'Full Name', 'Email', 'Phone', 'Department',
                   'Designation', 'Status', 'Employment Type', 'Date Joined',
                   'Basic Salary', 'City', 'State']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = font
            cell.fill = fill
            cell.alignment = align
            ws.column_dimensions[cell.column_letter].width = 18

        employees = Employee.objects.select_related('department', 'designation').all()
        for row, emp in enumerate(employees, 2):
            ws.cell(row=row, column=1, value=emp.employee_id)
            ws.cell(row=row, column=2, value=emp.full_name)
            ws.cell(row=row, column=3, value=emp.email)
            ws.cell(row=row, column=4, value=emp.phone)
            ws.cell(row=row, column=5, value=emp.department.name if emp.department else '')
            ws.cell(row=row, column=6, value=emp.designation.title if emp.designation else '')
            ws.cell(row=row, column=7, value=emp.get_status_display())
            ws.cell(row=row, column=8, value=emp.get_employment_type_display())
            ws.cell(row=row, column=9, value=str(emp.date_joined))
            ws.cell(row=row, column=10, value=float(emp.basic_salary))
            ws.cell(row=row, column=11, value=emp.city)
            ws.cell(row=row, column=12, value=emp.state)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="employees_{datetime.date.today()}.xlsx"'
        wb.save(response)
        return response


class SalaryReportView(APIView):
    permission_classes = [IsAuthenticated, IsHR]

    def get(self, request):
        month = request.query_params.get('month')
        year = request.query_params.get('year')
        employee_id = request.query_params.get('employee')

        qs = SalaryRecord.objects.select_related('employee', 'employee__department').all()
        if month:
            qs = qs.filter(month=month)
        if year:
            qs = qs.filter(year=year)
        if employee_id:
            qs = qs.filter(employee_id=employee_id)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Salary Report"

        font, fill, align = make_header_style()
        headers = ['Emp ID', 'Name', 'Department', 'Month', 'Year',
                   'Basic', 'HRA', 'Allowances', 'PF Deduction',
                   'Tax', 'Other Deductions', 'Net Salary', 'Status', 'Paid Date']

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = font
            cell.fill = fill
            cell.alignment = align
            ws.column_dimensions[cell.column_letter].width = 16

        for row, rec in enumerate(qs, 2):
            ws.cell(row=row, column=1, value=rec.employee.employee_id)
            ws.cell(row=row, column=2, value=rec.employee.full_name)
            ws.cell(row=row, column=3, value=rec.employee.department.name if rec.employee.department else '')
            ws.cell(row=row, column=4, value=rec.get_month_display())
            ws.cell(row=row, column=5, value=rec.year)
            ws.cell(row=row, column=6, value=float(rec.basic_salary))
            ws.cell(row=row, column=7, value=float(rec.hra))
            ws.cell(row=row, column=8, value=float(rec.allowances))
            ws.cell(row=row, column=9, value=float(rec.pf_deduction))
            ws.cell(row=row, column=10, value=float(rec.tax_deduction))
            ws.cell(row=row, column=11, value=float(rec.deductions))
            ws.cell(row=row, column=12, value=float(rec.net_salary))
            ws.cell(row=row, column=13, value=rec.get_status_display())
            ws.cell(row=row, column=14, value=str(rec.paid_date) if rec.paid_date else '')

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        fname = f"salary_report_{month or 'all'}_{year or 'all'}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{fname}"'
        wb.save(response)
        return response


class AssetReportView(APIView):
    permission_classes = [IsAuthenticated, IsHR]

    def get(self, request):
        qs = Asset.objects.select_related('employee').all()
        status_filter = request.query_params.get('status')
        if status_filter:
            qs = qs.filter(status=status_filter)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Asset Report"

        font, fill, align = make_header_style()
        headers = ['Emp ID', 'Employee Name', 'Asset Type', 'Status',
                   'Serial Number', 'Model', 'Issued Date', 'Returned Date', 'Notes']

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = font
            cell.fill = fill
            cell.alignment = align
            ws.column_dimensions[cell.column_letter].width = 18

        for row, asset in enumerate(qs, 2):
            ws.cell(row=row, column=1, value=asset.employee.employee_id)
            ws.cell(row=row, column=2, value=asset.employee.full_name)
            ws.cell(row=row, column=3, value=asset.get_asset_type_display())
            ws.cell(row=row, column=4, value=asset.get_status_display())
            ws.cell(row=row, column=5, value=asset.serial_number)
            ws.cell(row=row, column=6, value=asset.model_name)
            ws.cell(row=row, column=7, value=str(asset.issued_date) if asset.issued_date else '')
            ws.cell(row=row, column=8, value=str(asset.returned_date) if asset.returned_date else '')
            ws.cell(row=row, column=9, value=asset.notes)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="assets_{datetime.date.today()}.xlsx"'
        wb.save(response)
        return response


class DashboardStatsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        today = datetime.date.today()
        emp_qs = Employee.objects
        salary_qs = SalaryRecord.objects.filter(month=today.month, year=today.year)
        return Response({
            'employees': {
                'total': emp_qs.count(),
                'active': emp_qs.filter(status='active').count(),
                'resigned': emp_qs.filter(status='resigned').count(),
                'on_leave': emp_qs.filter(status='on_leave').count(),
                'this_month_joined': emp_qs.filter(
                    date_joined__month=today.month,
                    date_joined__year=today.year
                ).count(),
            },
            'salary': {
                'paid': salary_qs.filter(status='paid').count(),
                'unpaid': salary_qs.filter(status='unpaid').count(),
                'on_hold': salary_qs.filter(status='on_hold').count(),
            },
            'assets': {
                'laptops_received': Asset.objects.filter(asset_type='laptop', status='received').count(),
                'laptops_pending': Asset.objects.filter(asset_type='laptop', status__in=['not_received','pending']).count(),
                'id_cards_received': Asset.objects.filter(asset_type='id_card', status='received').count(),
                'id_cards_pending': Asset.objects.filter(asset_type='id_card', status__in=['not_received','pending']).count(),
                'kits_received': Asset.objects.filter(asset_type='kit', status='received').count(),
                'kits_pending': Asset.objects.filter(asset_type='kit', status__in=['not_received','pending']).count(),
            }
        })
